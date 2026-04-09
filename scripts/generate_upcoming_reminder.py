from __future__ import annotations

import argparse
import csv
from datetime import date, datetime, timedelta
from pathlib import Path


DEFAULT_INPUT = "data/events.csv"
DEFAULT_OUTPUT = "reminders/upcoming.md"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a reminder file for upcoming date/name entries."
    )
    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT,
        help="Path to the source CSV or Excel file.",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help="Path to the generated markdown file.",
    )
    parser.add_argument(
        "--date-column",
        default="date",
        help="Column containing the target date.",
    )
    parser.add_argument(
        "--name-column",
        default="name",
        help="Column containing the name to mention.",
    )
    parser.add_argument(
        "--lookahead-days",
        type=int,
        default=7,
        help="Include rows whose dates fall within this many days from today.",
    )
    parser.add_argument(
        "--today",
        default=None,
        help="Override today's date in YYYY-MM-DD format for testing.",
    )
    return parser.parse_args()


def load_table(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8") as handle:
            return list(csv.DictReader(handle))
    if path.suffix.lower() in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Excel support requires openpyxl. Install it or use a CSV file."
            ) from exc

        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        output: list[dict[str, str]] = []

        for row in rows[1:]:
            entry: dict[str, str] = {}
            for header, value in zip(headers, row):
                entry[header] = "" if value is None else str(value)
            output.append(entry)

        return output
    raise ValueError(f"Unsupported file type: {path.suffix}")


def parse_date(value: str) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def build_output(upcoming: list[dict[str, str | int | date]], start: date, end: date) -> str:
    lines = [
        "# Upcoming reminders",
        "",
        f"Generated for {start} through {end}.",
        "",
    ]

    if not upcoming:
        lines.append("No upcoming dates in this window.")
        lines.append("")
        return "\n".join(lines)

    lines.append("| Date | Name | Days remaining |")
    lines.append("| --- | --- | ---: |")

    for row in upcoming:
        lines.append(
            f"| {row['target_date']} | {row['person_name']} | {row['days_remaining']} |"
        )

    lines.append("")
    lines.append("Please review the upcoming names and dates before merging.")
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    rows = load_table(input_path)
    if rows:
        available_columns = set(rows[0].keys())
        required_columns = {args.date_column, args.name_column}
        missing = required_columns.difference(available_columns)
        if missing:
            missing_text = ", ".join(sorted(missing))
            raise KeyError(f"Missing required columns: {missing_text}")

    today = (
        datetime.strptime(args.today, "%Y-%m-%d").date()
        if args.today
        else date.today()
    )
    window_end = today + timedelta(days=args.lookahead_days)

    upcoming: list[dict[str, str | int | date]] = []
    for row in rows:
        target_date = parse_date(row.get(args.date_column, ""))
        person_name = str(row.get(args.name_column, "")).strip()
        if target_date is None or not person_name:
            continue

        days_remaining = (target_date - today).days
        if today <= target_date <= window_end:
            upcoming.append(
                {
                    "target_date": target_date,
                    "person_name": person_name,
                    "days_remaining": days_remaining,
                }
            )

    upcoming.sort(key=lambda row: (row["target_date"], row["person_name"]))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(build_output(upcoming, today, window_end), encoding="utf-8")


if __name__ == "__main__":
    main()
